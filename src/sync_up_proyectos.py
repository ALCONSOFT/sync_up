import pyodbc
import sys
import xmlrpc.client
import os
from datetime import datetime
from openpyxl import Workbook
import xlsxwriter
from datetime import datetime

# ODOO 14

url = "http://odoradita.com:8069"
db = "test2_CADASA_data"
username = 'soporte@alconsoft.net'
password = "2010Sistech"
max_registros = 501

#Para DOS/Windows
os.system ("cls")
print("INICIANDO RUTINA DE SINCRONIZACION DE PROYECTOS")
common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
print("common version: ")
print(common.version())

#User Identifier
uid = common.authenticate(db, username, password, {})
print("uid: ",uid)

# Calliing methods
models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
models.execute_kw(db, uid, password,
              'res.partner', 'check_access_rights',
              ['read'], {'raise_exception': False})

#Lectura de registros de PRueba de Departamento
models_dep = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
models_dep.execute_kw(db, uid, password,
                     'hr.department', 'check_access_rights',
                     ['read'], {'raise_exception': False})
    
filtro = [[ ['company_id', '=', 1],['active','=',1]]]  #lista de python
#registros = models_dep.execute_kw(db, uid, password, 'hr.department', 'search_count', filtro)
ids = models_dep.execute_kw(db, uid, password, 'hr.department', 'search_read',
    filtro, {'fields':['name', 'manager_id'],'limit': max_registros})
if ids == 0:
    print("Sin registros")
else:
    print("Se encontraron registros:", ids)
    #[regs] = models_dep.execute_kw(db, uid, password,
    #    'hr.department', 'read', [ids], {'fields':['name', 'manager_id']})
    print("Cantidad de Registros: ", len(ids))
    print("Tipo de Dato:", type([ids]))
    for elemento in ids:
        print("Elemento:", elemento)
    print("########## FIN DE PRUEBA DE DEPARTAMENTOS ############")
# - Lectura de registros de Proyectos: DATA MASTER I
models_proj = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
models_proj.execute_kw(db, uid, password,
                     'project.project', 'check_access_rights',
                     ['read'], {'raise_exception': False})
    
filtro = [[ ['company_id', '=', 1],['active','=',1]]]  #lista de python
#ids = models_dep.execute_kw(db, uid, password, 'project.project', 'search_read',
#    filtro, {'limit': max_registros})
ids = models_dep.execute_kw(db, uid, password, 'project.project', 'search_read',
    filtro, {'fields':['id','name','description','active','partner_id','company_id','fincas_pma',
    'zafra','odc','frente','subfinca','up','lote','has','correg','dist',
    'desr','ubic','tds','fecha_est_cosecha','tipocorte','variedad','fdc',
    'fds','hdc','hdq','hasq','tonq','hasv','tonv','tche1','tche2','tche3',
    'hasc','toncos','tonme','tonrt','tchr','difton','difprc','are','bx',
    'sac','pza','red','ph','tch_01','dif_01','tch_02','dif_02','tch_03',
    'dif_03','tch_04','dif_04','dosm','mad','fdam'],'limit': max_registros})
if ids == 0:
    print("Sin registros en Proyectos")
else:
    print("Se encontraron registros:", ids)
    print("Cantidad de Registros: ", len(ids))
    print("Tipo de Dato:", type([ids]))
    ############# INICIALINZADO PARAMETROS PARA EXCEL
    wb = Workbook()
    ruta = 'salida.xlsx'
    hoja = wb.active
    hoja.title = "Lista de Proyectos - UPLotes"
    
    fila = 1 #Fila donde empezamos
    col_id = 1 #Columna donde guardamos las fechas
    col_name = 2 #Columna donde guardamos el dato asociados a cada fecha
    col_up = 3
    col_lote = 4
    fila=1
    for elemento in ids:
        #print("Tipo: ", type(elemento))
        print("Procesndo Fila: ",fila)
        #print("Elemento:", elemento)
        nombre_cols = elemento.keys()
        valores = elemento.values()
        if fila == 1:
            print("Imprimiendo Nombre de Columnas", nombre_cols)
            for val_col in nombre_cols:
                hoja.cell(column=col_id, row=1, value=val_col)
                col_id+=1
        else:
            print("Imprimiendo Valores de celda...", valores )
            print("Tipo variable <valores>:",type(valores))
            col_id = 1

            for val_celda in valores:
                print("Tipo variable <value>:",type(val_celda))
                if type(val_celda) is not list:
                    hoja.cell(column=col_id, row=fila, value=val_celda)
                    col_id+=1
                else:
                    print("Lista Ignorada: ", val_celda)
                    if val_celda != []:
                        print("val_celda:", val_celda[1])
                        hoja.cell(column=col_id, row=fila, value=val_celda[1])
                    else:
                        print("valor de lista en celda Ignorado por estar vacio!!!")
                    col_id+=1

        fila+=1

    print("########## FIN DE IMPRESION DE PROYECTOS ############")
    wb.save(filename = ruta)
############ fin de programa sincronizador ##########